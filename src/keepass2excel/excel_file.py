from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.fills import PatternFill

def index2letter(index):
    return chr(65 + index)  # 0 is 'A', 1 is 'B', ...

class ExcelFile():
    def __init__(self):
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.page_setup.orientation = "landscape"
        self.worksheet.page_margins.left = 0.6
        self.worksheet.page_margins.right = 0.6
        self.worksheet.page_margins.top = 0.6
        self.worksheet.page_margins.bottom = 0.6
        self.worksheet.print_title_rows = "1:1"  # first row on every page
        self.worksheet.oddFooter.left.text = "Page &[Page] of &N"
        self.worksheet.oddFooter.left.size = 9
        self.worksheet.oddFooter.left.font = "Tahoma,Italic"
        self.worksheet.oddFooter.left.color = "0000FF"

    def set_column_widths(self, col_widths):
        for i in range(0, len(col_widths)):
            self.worksheet.column_dimensions[index2letter(i)].width = col_widths[i]

    def set_cell(self, row, col, text, font="Calibri", size=12, foreground="000000", bold=False,
                background=None, halign="left", valign="top", wrap=True):
        cell_location = f"{index2letter(col)}{row}"
        cell = self.worksheet[cell_location]
        cell.value = text
        cell.font = Font(name=font, size=size, color=foreground, bold=bold)
        cell.alignment = Alignment(wrap_text=wrap, horizontal=halign, vertical=valign)
        if background:
            cell.fill = PatternFill("solid", start_color=background)
        cell.border = Border(top=Side(border_style="hair", color="AAAAAA"))


        # self.current_row = self.current_row + 1
        # cell_name_iter = ExcelColumnCellIterator(self.current_row)
        # if col_defs:
        #     for col_def in col_defs:
        #         if type(col_def) == str or type(col_def) == int:
        #             value = col_def
        #             styles = []
        #         else:
        #             col_def_as_ary = [ c for c in col_def ]
        #             value = col_def_as_ary.pop(0)
        #             styles = col_def_as_ary
        #         if type(value) == str:
        #             value = value.replace("{row}", str(self.current_row))
        #         cell = self.worksheet[cell_name_iter.next_cell_name()]
        #         cell.value = value
        #         for style in styles:
        #             if type(style) == Font:
        #                 cell.font = style
        #             elif type(style) == PatternFill:
        #                 cell.fill = style
        #             elif type(style) == Border:
        #                 cell.border = style
        #             elif type(style) == Alignment:
        #                 cell.alignment = style
        #             elif type(style) == str:
        #                 cell.number_format = style


    def save(self, out_stream):
        self.workbook.save(out_stream)
